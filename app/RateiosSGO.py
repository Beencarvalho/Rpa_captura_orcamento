from tqdm import tqdm
import duckdb
import requests
import sys
import time
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from util.api_token import api_budget, api_budget_months, headers

def show_startup_animation():
    # Desenho simples em ASCII
    logo = [
        "  #####   #####    ##### ",
        " #     # #     #  #     #",
        " #       #        #     #",
        "  #####  #  ####  #     #",
        "       # #     #  #     #",
        " #     # #     #  #     #",
        "  #####   #####    ##### "
    ]

    # Animação do desenho
    for line in logo:
        print(line)
        time.sleep(0.1)  # Pequeno delay para criar o efeito de "desenho"

    # Mensagem de inicialização
    print("\n\nIniciando conexão com API SGO")
    
    # Animação de carregamento
    loading_animation = ["[=     ]", "[==    ]", "[===   ]", "[====  ]", "[===== ]", "[======]"]
    for i in range(3):  # Repetir a animação algumas vezes
        for frame in loading_animation:
            sys.stdout.write("\r" + frame)
            sys.stdout.flush()
            time.sleep(0.2)  # Delay entre os frames
    print("\n\nConexão estabelecida com sucesso!")
    print("\n\nIniciando geração de arquivos:")

# Chamar a função para exibir a animação
show_startup_animation()

# Comando para obter o caminho padrão da area de trabalho em qualquer maquina
desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')

pasta_arquivos = os.path.join(desktop_path, 'Arquivos SGO')

# Certifique-se de que a pasta 'dados' exista
if not os.path.exists(pasta_arquivos):
    os.makedirs(pasta_arquivos)

# Definindo o caminho para salvar os arquivos
nome_arquivo1 = 'Validacao dos Dados SGO.xlsx'
nome_arquivo2 = 'Controladoria.xlsx'
file_path_geral = os.path.join(pasta_arquivos, nome_arquivo1)
file_path_grupo = os.path.join(pasta_arquivos, nome_arquivo2)

# Requisições para obter os dados de budget na API /budgets/get-all
# E as tratativas caso algum erro ocorra na requsição 
try:
    response = requests.get(api_budget, headers=headers)
    response.raise_for_status()  # Lança uma exceção se a resposta não for 2xx
    # Caso o status seja 200, processa o JSON normalmente
    budget = response.json()

except requests.exceptions.HTTPError as http_err:
    # Trata erros HTTP específicos com base no código de status
    if response.status_code == 401:
        print("Erro de autenticação. Verifique o token de acesso.")
    elif response.status_code == 404:
        print("Recurso não encontrado. Verifique a URL da API.")
    elif response.status_code == 500:
        print("Erro interno do servidor. Tente novamente mais tarde.")
    else:
        print(f"Erro HTTP ao acessar a API: {response.status_code} - {http_err}")
    sys.exit(1)  # Encerra o programa com código de erro

except requests.exceptions.RequestException as req_err:
    # Trata erros de conexão, tempo de espera, etc.
    print(f"Erro ao tentar se conectar à API: {req_err}")
    sys.exit(1)  # Encerra o programa com código de erro

except Exception as err:
    # Trata qualquer outro erro inesperado
    print(f"Ocorreu um erro inesperado: {err}")
    sys.exit(1)  # Encerra o programa com código de erro

# Criando um banco de dados temporário no DuckDB
con = duckdb.connect(database=':memory:')

# Convertendo os dados JSON para DataFrames do Pandas
budget_df = pd.json_normalize(budget, sep='_', meta=[
    'active', 'id', 'contractNumber', 'adjustmentMonth', 'adjustmentPercentage',
    'value', 'cycleId', 'budgetAccountId', 'supplierId', 'originId',
    'levelSixId', 'managerId', 'apportionmentId'
], record_prefix='', errors='ignore')

# Criando um DataFrame para armazenar os detalhes dos meses
budget_months_list = []

# Definindo o número máximo de tentativas para o caso de 429
max_retries = 3

# Barra de progresso com tqdm - OBTENDO DADOS DA API POR ID
with tqdm(total=len(budget), desc="Processando Orçamentos") as pbar:
    for budget_entry in budget:
        budget_id = budget_entry['id']
        retries = 0

        while retries < max_retries:
            response_months = requests.get(f"{api_budget_months}?budgetId={budget_id}", headers=headers)

            if response_months.status_code == 200:
                budget_months = response_months.json()
                budget_months_list.extend(budget_months)
                tqdm.write(f"Itens do orçamento do ID: {budget_id}, obtidos com sucesso.")
                time.sleep(1)  # Pequeno atraso após sucesso
                break

            elif response_months.status_code == 429:
                retries += 1
                tqdm.write(
                    f"Tempo limite da API excedido para o budgetId {budget_id}. "
                    f"Tentativa {retries} de {max_retries}. Aguardando antes de tentar novamente."
                )
                time.sleep(2 * retries)  # Tempo de espera exponencial

            else:
                tqdm.write(f"Erro ao obter os detalhes do orçamento para o budgetId {budget_id}: {response_months.status_code}")
                sys.exit(1)  # Encerra o programa se ocorrer outro erro
        else:
            tqdm.write(f"Falha ao obter os dados para o orçamento {budget_id} após {max_retries} tentativas.")

        pbar.update(1)  # Atualiza a barra de progresso após o término do processamento de cada orçamento

tqdm.write('\n\nDados obtidos, construindo arquivos...')

# Convertendo os dados de budget_months para um DataFrame
budget_months_df = pd.json_normalize(budget_months_list, sep='_', errors='ignore')

# Carregando os DataFrames para tabelas no DuckDB
con.register('budget', budget_df)
con.register('budget_months', budget_months_df)

# Unindo os dados de budget e budget months
df_geral = con.execute('''
    SELECT 
        b.id AS Id_Orçamento,
        b.contractNumber AS Contrato,
        b.adjustmentMonth AS Mes_Reajuste,
        b.adjustmentPercentage AS Reajuste_Percentual,
        b.value AS Valor,
        b.supplier_code AS Cod_Fornecedor,
        b.budgetAccount_code AS COD_CONTA_CONTABIL,
        b.budgetAccount_description AS DESC_CONTA_CONTABIL,
        b.supplier_description AS Fornecedor,
        b.origin_description AS Origem,
        bm.budgetApportionmentItem_sector_code AS COD_SETOR,
        bm.budgetApportionmentItem_sector_codeCostCenter AS COD_CCUSTO,
        bm.budgetApportionmentItem_sector_name AS CENTRO_CUSTO,
        bm.budgetApportionmentItem_base AS BASE,
        bm.budgetApportionmentItem_sector_company_name AS EMPRESA,
        b.levelSix_description AS Nivel,
        b.manager_description AS Gestor,
        b.apportionment_name AS Criterio,
        b.apportionment_description AS Descricao_criterio,
        b.cycle_budgetYear AS Ano,
        bm.january AS Janeiro,
        bm.february AS Fevereiro,
        bm.march AS Março,
        bm.april AS Abril,
        bm.may AS Maio,
        bm.june AS Junho,
        bm.july AS Julho,
        bm.august AS Agosto,
        bm.september AS Setembro,
        bm.october AS Outubro,
        bm.november AS Novembro,
        bm.december AS Dezembro,
        -- Soma total anual
        COALESCE(bm.january, 0) + COALESCE(bm.february, 0) + COALESCE(bm.march, 0) +
        COALESCE(bm.april, 0) + COALESCE(bm.may, 0) + COALESCE(bm.june, 0) +
        COALESCE(bm.july, 0) + COALESCE(bm.august, 0) + COALESCE(bm.september, 0) +
        COALESCE(bm.october, 0) + COALESCE(bm.november, 0) + COALESCE(bm.december, 0) AS Total_Anual
    FROM budget b
    JOIN budget_months bm ON b.id = bm.budgetId
''').fetchdf()

# Diretório para salvar os arquivos
output_dir = os.path.join(os.path.expanduser("~"), "Desktop", "Arquivos_Contratos")
os.makedirs(output_dir, exist_ok=True)

# Obtém a lista de budget IDs únicos
budget_ids = df_geral["Id_Orçamento"].dropna().unique()

# Caminho para salvar os arquivos
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
output_folder = os.path.join(desktop_path, "Arquivos SGO")

# Certifique-se de que a pasta exista
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

for budget_id in budget_ids:
    # Filtra o DataFrame para o budget ID atual
    df_budget = df_geral[df_geral["Id_Orçamento"] == budget_id]
    
    # Verificação se o DataFrame está vazio
    if df_budget.empty:
        print(f"Atenção: Não foram encontrados dados para o Budget ID {budget_id}. Pulando para o próximo budget.")
        continue  # Salta para o próximo budget no loop

    # Obter dados principais do budget para o cabeçalho
    criterio = df_budget["Criterio"].iloc[0]
    cod_conta_contabil = df_budget["COD_CONTA_CONTABIL"].iloc[0]
    desc_conta_contabil = df_budget["DESC_CONTA_CONTABIL"].iloc[0]
    fornecedor = df_budget["Fornecedor"].iloc[0]
    reajuste_percentual = df_budget["Reajuste_Percentual"].iloc[0]
    mes_reajuste = df_budget["Mes_Reajuste"].iloc[0]
    descricao_criterio = df_budget["Descricao_criterio"].iloc[0]

    # Nome do arquivo personalizado
    safe_criterio = criterio.replace("/", "_").replace("\\", "_").replace(" ", "_")
    safe_fornecedor = fornecedor.replace("/", "_").replace("\\", "_").replace(" ", "_") #separar só a primeiro nome
    file_name = f"2025_{safe_fornecedor}_{budget_id}.xlsx"
    file_path = os.path.join(output_folder, file_name)

    # Cálculo do percentual baseado na BASE
    total_base = df_budget["BASE"].sum()
    df_budget = df_budget.copy()
    df_budget["Percentual"] = (
        (df_budget["BASE"] / total_base) * 100 if total_base != 0 else 0
    )

    # Criando o Excel
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        # Criando o workbook e aba principal com título único
        workbook = writer.book
        aba_title = safe_fornecedor
        worksheet = workbook.create_sheet(title=aba_title)

        # Remover aba padrão criada automaticamente
        if "Sheet" in workbook.sheetnames:
            del workbook["Sheet"]

        # Adiciona o cabeçalho (linhas 1-4)
        worksheet.merge_cells("A1:C1")
        worksheet.merge_cells("A2:C2")
        worksheet["A1"] = "Critério"
        worksheet["A2"] = criterio

        worksheet["D1"] = "Conta Contábil"
        worksheet["D2"] = cod_conta_contabil

        worksheet.merge_cells("E1:F1")
        worksheet.merge_cells("E2:F2")
        worksheet["E1"] = "Descrição de Conta"
        worksheet["E2"] = desc_conta_contabil

        # Linha 3
        worksheet.append(["NR_CONTRATO", "CD_FORNECEDOR", "NM_FORNECEDOR", "Mês Reajuste", "% de Reajuste", "Regra"])
        worksheet.append([
            "", fornecedor, "", mes_reajuste, f"{reajuste_percentual:.2f}%", descricao_criterio
        ])

        # Deixe as linhas 5 a 8 vazias
        for _ in range(2):
            worksheet.append([])

        # Adiciona o cabeçalho da tabela principal (a partir da linha 6)
        header = ["EMPRESA", "COD_SETOR", "COD_CCUSTO", "CENTRO_CUSTO", "BASE", "PERCENTUAL",
                  "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
                  "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO", "TOTAL_ANUAL"]
        worksheet.append(header)

        # Adiciona os dados da tabela principal
        for _, row in df_budget.iterrows():
            worksheet.append([
                row["EMPRESA"], row["COD_SETOR"], row["COD_CCUSTO"], row["CENTRO_CUSTO"],
                row["BASE"], f"{row['Percentual']:.2f}%",
                row["Janeiro"], row["Fevereiro"], row["Março"], row["Abril"], row["Maio"], row["Junho"],
                row["Julho"], row["Agosto"], row["Setembro"], row["Outubro"], row["Novembro"], row["Dezembro"],
                row["Total_Anual"]
            ])

    print(f"Arquivo gerado para Budget ID {budget_id}: {file_path}")

print("Processamento concluído.")
